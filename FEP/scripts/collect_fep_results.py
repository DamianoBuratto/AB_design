#!/usr/bin/env python3
"""
Collect FEP ΔΔG results into a summary table.

Scans  FEP/outputs/replica{1..N}/<design>/ddG_results.txt  and assembles
the values into a CSV (and optionally Excel) table matching the format:

    design_group | Design | ΔΔG (kcal/mol) | Error (kcal/mol) | ΔΔG (kJ/mol) | Error (kJ/mol)

Designs whose analysis is not yet finished appear with empty value columns
and a "PENDING" status note.

Usage
-----
    # Collect all designs found under FEP/outputs/
    python scripts/collect_fep_results.py

    # Specify a custom outputs root or output file
    python scripts/collect_fep_results.py --outputs /path/to/FEP/outputs \
                                           --out fep_summary.csv

    # Filter to specific designs (names must match directory names exactly)
    python scripts/collect_fep_results.py --designs design_454_dldesign_3_best \
                                                     design_373_dldesign_13_best

    # Read design names from a text file (one per line)
    python scripts/collect_fep_results.py --designs-file my_designs.txt

    # Also save as Excel (.xlsx) — requires openpyxl
    python scripts/collect_fep_results.py --excel
"""

import argparse
import os
import re
import sys
from pathlib import Path

# ── regex to parse the final ΔΔG line in ddG_results.txt ─────────────────────
# Format written by analyze_fep.py:
#   ΔΔG_binding (final) = +4.717 ± 0.800 kJ/mol  (+1.127 ± 0.191 kcal/mol)
_PATTERN = re.compile(
    r'ΔΔG_binding\s*\(final\)\s*=\s*'
    r'([+\-]?\d+\.\d+)\s*±\s*(\d+\.\d+)\s*kJ/mol'
    r'\s+\(([+\-]?\d+\.\d+)\s*±\s*(\d+\.\d+)\s*kcal/mol\)'
)

# Fallback: some older versions wrote the line without the kcal section
_PATTERN_KJONLY = re.compile(
    r'ΔΔG_binding\s*\(final\)\s*=\s*'
    r'([+\-]?\d+\.\d+)\s*±\s*(\d+\.\d+)\s*kJ/mol'
)


def parse_ddg_file(filepath: Path):
    """Return (ddG_kcal, err_kcal, ddG_kJ, err_kJ) or None if not parseable."""
    try:
        text = filepath.read_text(encoding='utf-8')
    except OSError:
        return None

    m = _PATTERN.search(text)
    if m:
        ddG_kJ  = float(m.group(1))
        err_kJ  = float(m.group(2))
        ddG_kc  = float(m.group(3))
        err_kc  = float(m.group(4))
        return ddG_kc, err_kc, ddG_kJ, err_kJ

    # fallback: kJ/mol only — convert to kcal/mol
    m2 = _PATTERN_KJONLY.search(text)
    if m2:
        ddG_kJ = float(m2.group(1))
        err_kJ = float(m2.group(2))
        return ddG_kJ / 4.184, err_kJ / 4.184, ddG_kJ, err_kJ

    return None


def scan_outputs(outputs_root: Path, n_replicas: int = 3):
    """Yield (design_group, design_name, result_or_None) for every design dir found.

    design_group : e.g. "replica1"
    design_name  : e.g. "design_454_dldesign_3_best"
    result       : (ddG_kcal, err_kcal, ddG_kJ, err_kJ) or None
    """
    for rep_n in range(1, n_replicas + 1):
        rep_dir = outputs_root / f"replica{rep_n}"
        if not rep_dir.is_dir():
            continue
        for design_dir in sorted(rep_dir.iterdir()):
            if not design_dir.is_dir():
                continue
            result_file = design_dir / 'ddG_results.txt'
            result = parse_ddg_file(result_file) if result_file.is_file() else None
            yield f"replica{rep_n}", design_dir.name, result


def collect(outputs_root: Path, design_filter: set | None, n_replicas: int = 3):
    """Return list of row dicts for the summary table."""
    rows = []
    for group, name, result in scan_outputs(outputs_root, n_replicas):
        if design_filter and name not in design_filter:
            continue
        if result is not None:
            ddG_kc, err_kc, ddG_kJ, err_kJ = result
            rows.append({
                'design_group':    group,
                'Design':          name,
                'ΔΔG (kcal/mol)':  round(ddG_kc, 3),
                'Error (kcal/mol)': f'±{err_kc:.3f}',
                'ΔΔG (kJ/mol)':    round(ddG_kJ, 3),
                'Error (kJ/mol)':  f'±{err_kJ:.3f}',
                '_status':         'done',
            })
        else:
            rows.append({
                'design_group':    group,
                'Design':          name,
                'ΔΔG (kcal/mol)':  '',
                'Error (kcal/mol)': '',
                'ΔΔG (kJ/mol)':    '',
                'Error (kJ/mol)':  '',
                '_status':         'PENDING',
            })
    return rows


def write_csv(rows: list, outpath: Path, verbose: bool = True):
    import csv
    cols = ['design_group', 'Design',
            'ΔΔG (kcal/mol)', 'Error (kcal/mol)',
            'ΔΔG (kJ/mol)',   'Error (kJ/mol)']
    outpath.parent.mkdir(parents=True, exist_ok=True)
    with open(outpath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=cols, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(rows)
    if verbose:
        print(f"  CSV  → {outpath}")


def write_excel(rows: list, outpath: Path, verbose: bool = True):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("WARNING: openpyxl not installed; skipping Excel output.  "
              "Install with:  pip install openpyxl")
        return

    cols = ['design_group', 'Design',
            'ΔΔG (kcal/mol)', 'Error (kcal/mol)',
            'ΔΔG (kJ/mol)',   'Error (kJ/mol)']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FEP Results"

    # Header style
    header_fill = PatternFill("solid", fgColor="2E86AB")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_side   = Side(style='thin', color='CCCCCC')
    thin_border = Border(left=thin_side, right=thin_side,
                         top=thin_side,  bottom=thin_side)

    for col_i, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_i, value=col_name)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Data rows
    pending_fill = PatternFill("solid", fgColor="FFF3CD")   # pale yellow
    done_fill    = PatternFill("solid", fgColor="FFFFFF")

    for row_i, row in enumerate(rows, start=2):
        fill = pending_fill if row.get('_status') == 'PENDING' else done_fill
        for col_i, col_name in enumerate(cols, start=1):
            val  = row.get(col_name, '')
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.fill   = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    # Auto column width
    for col_i, col_name in enumerate(cols, start=1):
        max_len = max(
            len(str(col_name)),
            *[len(str(r.get(col_name, ''))) for r in rows],
        )
        ws.column_dimensions[get_column_letter(col_i)].width = max_len + 4

    ws.freeze_panes = 'A2'
    outpath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(outpath)
    if verbose:
        print(f"  XLSX → {outpath}")


def print_table(rows: list):
    """Print a human-readable summary to stdout."""
    col_w = {
        'design_group':    12,
        'Design':          35,
        'ΔΔG (kcal/mol)':  14,
        'Error (kcal/mol)': 16,
        'ΔΔG (kJ/mol)':    13,
        'Error (kJ/mol)':  13,
        '_status':         8,
    }
    header = (
        f"{'design_group':<{col_w['design_group']}}  "
        f"{'Design':<{col_w['Design']}}  "
        f"{'ΔΔG (kcal/mol)':>{col_w['ΔΔG (kcal/mol)']}}  "
        f"{'Error (kcal/mol)':>{col_w['Error (kcal/mol)']}}  "
        f"{'ΔΔG (kJ/mol)':>{col_w['ΔΔG (kJ/mol)']}}  "
        f"{'Error (kJ/mol)':>{col_w['Error (kJ/mol)']}}  "
        f"{'status':<{col_w['_status']}}"
    )
    print("\n" + "─" * len(header))
    print(header)
    print("─" * len(header))
    n_done = n_pending = 0
    for r in rows:
        status = r.get('_status', '')
        print(
            f"{r['design_group']:<{col_w['design_group']}}  "
            f"{r['Design']:<{col_w['Design']}}  "
            f"{str(r['ΔΔG (kcal/mol)']):>{col_w['ΔΔG (kcal/mol)']}}  "
            f"{str(r['Error (kcal/mol)']):>{col_w['Error (kcal/mol)']}}  "
            f"{str(r['ΔΔG (kJ/mol)']):>{col_w['ΔΔG (kJ/mol)']}}  "
            f"{str(r['Error (kJ/mol)']):>{col_w['Error (kJ/mol)']}}  "
            f"{status:<{col_w['_status']}}"
        )
        if status == 'done':
            n_done += 1
        else:
            n_pending += 1
    print("─" * len(header))
    print(f"  Total: {len(rows)}  |  done: {n_done}  |  pending: {n_pending}\n")


# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Collect FEP ΔΔG results into a summary table.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        '--outputs', default=None,
        help="Path to FEP outputs root directory  "
             "(default: auto-detect as ~/FEP/outputs or ./outputs relative to this script)",
    )
    parser.add_argument(
        '--out', default=None,
        help="Output CSV file path  (default: <outputs_root>/../fep_summary.csv)",
    )
    parser.add_argument(
        '--excel', action='store_true',
        help="Also write an Excel (.xlsx) file alongside the CSV",
    )
    parser.add_argument(
        '--designs', nargs='+', metavar='DESIGN',
        help="Only include these design names (space-separated)",
    )
    parser.add_argument(
        '--designs-file', metavar='FILE',
        help="Read design names to include from a text file (one per line)",
    )
    parser.add_argument(
        '--replicas', type=int, default=3,
        help="Number of outer replica directories to scan (default: 3)",
    )
    parser.add_argument(
        '--no-print', action='store_true',
        help="Suppress the console table printout",
    )
    args = parser.parse_args()

    # ── Resolve outputs root ──────────────────────────────────────────────────
    if args.outputs:
        outputs_root = Path(args.outputs).expanduser().resolve()
    else:
        # Try common HPC path first, then fall back to path relative to this script
        hpc_path = Path.home() / "FEP" / "outputs"
        script_rel = Path(__file__).parent.parent / "outputs"
        if hpc_path.is_dir():
            outputs_root = hpc_path
        elif script_rel.is_dir():
            outputs_root = script_rel.resolve()
        else:
            print(
                "ERROR: Could not auto-detect outputs directory.\n"
                "  Pass --outputs /path/to/FEP/outputs  explicitly.",
                file=sys.stderr,
            )
            sys.exit(1)

    print(f"Scanning: {outputs_root}")
    if not outputs_root.is_dir():
        print(f"ERROR: outputs directory does not exist: {outputs_root}", file=sys.stderr)
        sys.exit(1)

    # ── Design filter ─────────────────────────────────────────────────────────
    design_filter: set | None = None
    names: list[str] = []
    if args.designs:
        names.extend(args.designs)
    if args.designs_file:
        fpath = Path(args.designs_file)
        if not fpath.is_file():
            print(f"ERROR: designs file not found: {fpath}", file=sys.stderr)
            sys.exit(1)
        names.extend(
            line.strip() for line in fpath.read_text().splitlines()
            if line.strip() and not line.startswith('#')
        )
    if names:
        design_filter = set(names)
        print(f"Filter: {len(design_filter)} design(s) specified")

    # ── Collect ───────────────────────────────────────────────────────────────
    rows = collect(outputs_root, design_filter, n_replicas=args.replicas)
    if not rows:
        print("No designs found.  Check --outputs path or --designs filter.")
        sys.exit(0)

    if not args.no_print:
        print_table(rows)

    # ── Write outputs ─────────────────────────────────────────────────────────
    if args.out:
        csv_path = Path(args.out).expanduser().resolve()
    else:
        csv_path = outputs_root.parent / "fep_summary.csv"

    write_csv(rows, csv_path)

    if args.excel:
        xlsx_path = csv_path.with_suffix('.xlsx')
        write_excel(rows, xlsx_path)

    print("Done.")


if __name__ == '__main__':
    main()
