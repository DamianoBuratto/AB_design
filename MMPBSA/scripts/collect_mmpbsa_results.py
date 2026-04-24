#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Collect MMPBSA results into a summary CSV table.

Scans  MMPBSA/results/replica{1..N}/<design>/mmpbsa_r{1..N}/FINAL_RESULTS_MMPBSA.dat
and extracts binding free-energy components into a clean CSV.

Directory structure:

    MMPBSA/results/
    └── replica1/                          <- outer (MD) replica
        └── design_373_dldesign_13_best/   <- design name
            └── mmpbsa_r1/                 <- inner (MMPBSA) replica
                ├── FINAL_RESULTS_MMPBSA.dat
                ├── gmx_MMPBSA_plot.csv
                ├── DONE
                └── ...

Output CSV columns:

    outer_replica  design  mmpbsa_replica  status
    COMPLEX_TOTAL_avg  COMPLEX_TOTAL_std  COMPLEX_TOTAL_stderr
    RECEPTOR_TOTAL_avg  ...
    LIGAND_TOTAL_avg  ...
    DELTA_TOTAL_avg  DELTA_TOTAL_std  DELTA_TOTAL_stderr
    DELTA_VDWAALS_avg  DELTA_VDWAALS_std
    DELTA_EEL_avg  DELTA_EEL_std
    DELTA_PB_avg  DELTA_PB_std
    DELTA_POLAR_avg  DELTA_POLAR_std
    DELTA_NONPOLAR_avg  DELTA_NONPOLAR_std
    DELTA_GAS_avg  DELTA_GAS_std
    DELTA_SOLV_avg  DELTA_SOLV_std

Optionally aggregates across MMPBSA replicas (mean ± SEM of DELTA TOTAL).

Usage:

    # Collect all designs found under ~/MMPBSA/results/
    python scripts/collect_mmpbsa_results.py

    # Custom paths
    python scripts/collect_mmpbsa_results.py \\
        --results /public/home/xuziyi/MMPBSA/results \\
        --out /public/home/xuziyi/MMPBSA/mmpbsa_summary.csv

    # Also save as Excel
    python scripts/collect_mmpbsa_results.py --excel

    # Filter to specific designs
    python scripts/collect_mmpbsa_results.py --designs design_404 design_373_dldesign_13_best

    # Include per-frame CSV data (gmx_MMPBSA_plot.csv)
    python scripts/collect_mmpbsa_results.py --per-frame

    # Only per-design averages (collapse MMPBSA replicas)
    python scripts/collect_mmpbsa_results.py --aggregate-replicas
"""

import argparse
import csv
import os
import re
import sys
from pathlib import Path
from collections import defaultdict

# ── regex patterns for parsing FINAL_RESULTS_MMPBSA.dat ────────────────────
#
# Actual file format (gmx_MMPBSA 1.6.4, Linear PB):
#
#   POISSON BOLTZMANN:
#
#   Complex:
#   Energy Component       Average     SD(Prop.)         SD   SEM(Prop.)        SEM
#   -------------------------------------------------------------------------------
#   BOND                    677.67         22.98      22.98         1.32       1.32
#   ...
#   TOTAL                 -4828.47        164.43      52.06         9.48       3.00
#
#   Receptor:
#   ...
#
#   Ligand:
#   ...
#
#   Delta (Complex - Receptor - Ligand):
#   Energy Component       Average     SD(Prop.)         SD   SEM(Prop.)        SEM
#   -------------------------------------------------------------------------------
#   ΔBOND                     0.00          4.29       0.00         0.25       0.00
#   ...
#   ΔTOTAL                   18.70         10.37      16.51         0.60       0.95
#
# Key: component names in Delta section use Unicode Δ (U+0394) prefix,
# and each row has 5 numeric columns: Average, SD(Prop.), SD, SEM(Prop.), SEM.

# Section headers in the output file
_SECTION_HEADERS = {
    'Complex:':                                    'COMPLEX',
    'Receptor:':                                  'RECEPTOR',
    'Ligand:':                                     'LIGAND',
    'Delta (Complex - Receptor - Ligand):':       'DELTA',
}

# Map raw component names (after stripping whitespace) to canonical keys.
# Delta section names include Unicode Δ prefix.
_DELTA_NAME_MAP = {
    '\u0394BOND':      'DELTA_BOND',
    '\u0394ANGLE':     'DELTA_ANGLE',
    '\u0394DIHED':    'DELTA_DIHED',
    '\u0394UB':        'DELTA_UB',
    '\u0394IMP':       'DELTA_IMP',
    '\u0394CMAP':      'DELTA_CMAP',
    '\u0394VDWAALS':  'DELTA_VDWAALS',
    '\u0394EEL':       'DELTA_EEL',
    '\u03941-4 VDW':  'DELTA_1-4_VDW',
    '\u03941-4 EEL':  'DELTA_1-4_EEL',
    '\u0394EPB':       'DELTA_PB',
    '\u0394ENPOLAR':  'DELTA_POLAR',
    '\u0394EDISPER':  'DELTA_NONPOLAR',
    '\u0394GGAS':      'DELTA_GAS',
    '\u0394GSOLV':     'DELTA_SOLV',
    '\u0394TOTAL':     'DELTA_TOTAL',
}

# For Complex/Receptor/Ligand sections, TOTAL is the main summary row
_SUMMARY_ROWS = {
    'GGAS':  'GAS',
    'GSOLV': 'SOLV',
    'TOTAL': 'TOTAL',
}


def _parse_energy_row(line: str) -> tuple[str, float, float, float, float, float] | None:
    """Parse a single energy row from FINAL_RESULTS_MMPBSA.dat.

    Expected format (5 numeric columns after component name):
        Component_Name   Average  SD(Prop.)  SD  SEM(Prop.)  SEM

    Returns (component_name, avg, sd_prop, sd, sem_prop, sem) or None.
    """
    # Strategy: find the last 5 float-like tokens (they are the numeric columns),
    # everything before is the component name.
    # Use \s{2,} as separator between name and first number to handle
    # multi-word names like "1-4 VDW" and "Δ1-4 EEL".
    m = re.match(
        r'^\s*(.+?)\s{2,}'
        r'([+-]?\d+\.\d+)\s+'
        r'([+-]?\d+\.\d+)\s+'
        r'([+-]?\d+\.\d+)\s+'
        r'([+-]?\d+\.\d+)\s+'
        r'([+-]?\d+\.\d+)\s*$',
        line,
    )
    if not m:
        return None
    name = m.group(1).strip()
    return (
        name,
        float(m.group(2)),   # Average
        float(m.group(3)),   # SD(Prop.)
        float(m.group(4)),   # SD
        float(m.group(5)),   # SEM(Prop.)
        float(m.group(6)),   # SEM
    )


def parse_final_results(filepath: Path) -> dict | None:
    """Parse FINAL_RESULTS_MMPBSA.dat and return a dict of energy components.

    Returns dict with keys like 'DELTA_TOTAL_avg', 'DELTA_TOTAL_std',
    'DELTA_TOTAL_stderr', 'COMPLEX_TOTAL_avg', etc., or None on failure.

    Numeric columns stored:
        _avg   = Average (column 1)
        _std   = SD      (column 3, sample standard deviation)
        _stderr = SEM    (column 5, sample standard error of the mean)
    """
    try:
        text = filepath.read_text(encoding='utf-8', errors='replace')
    except OSError:
        return None

    results = {}
    current_section = None

    for raw_line in text.splitlines():
        line = raw_line.strip()

        # Detect section headers
        for header_text, section_name in _SECTION_HEADERS.items():
            if line.startswith(header_text.rstrip(':')) and (':' in line or line == header_text.rstrip(':')):
                current_section = section_name
                break
        else:
            # Not a section header — try to parse as energy row
            parsed = _parse_energy_row(raw_line)
            if parsed is None:
                continue
            comp_name, avg, sd_prop, sd, sem_prop, sem = parsed

            if current_section == 'DELTA':
                # Delta section: component names have Unicode Δ prefix
                canonical = _DELTA_NAME_MAP.get(comp_name)
                if canonical is None:
                    # Fallback: try replacing Δ with DELTA_ prefix
                    if comp_name.startswith('\u0394'):
                        canonical = 'DELTA_' + comp_name[1:].replace(' ', '_').upper()
                    else:
                        # Also handle ASCII "DELTA" prefix
                        canonical = comp_name.replace(' ', '_').upper()
                results[f'{canonical}_avg'] = avg
                results[f'{canonical}_std'] = sd
                results[f'{canonical}_stderr'] = sem

            elif current_section in ('COMPLEX', 'RECEPTOR', 'LIGAND'):
                # Complex/Receptor/Ligand sections
                # Store summary rows (GGAS, GSOLV, TOTAL) and all others
                prefix = current_section
                canonical = _SUMMARY_ROWS.get(comp_name, comp_name.upper())
                key = f'{prefix}_{canonical}'
                results[f'{key}_avg'] = avg
                results[f'{key}_std'] = sd
                results[f'{key}_stderr'] = sem

    # Minimal validation: we need at least DELTA_TOTAL
    if 'DELTA_TOTAL_avg' not in results:
        return None

    return results


def parse_per_frame_csv(filepath: Path) -> dict | None:
    """Parse gmx_MMPBSA_plot.csv for per-frame DELTA TOTAL series.

    Returns dict with 'per_frame_delta_total' (list of floats) or None.
    The CSV has a header line starting with '# Time' or 'Time', then
    columns: Time, COM-BOND, ..., COM-TOTAL, ..., DELTA-TOTAL, ...
    """
    try:
        with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
            lines = f.readlines()
    except OSError:
        return None

    if len(lines) < 2:
        return None

    # Find DELTA TOTAL column index (may be "DELTA-TOTAL", "\u0394-TOTAL", or similar)
    header = lines[0].strip()
    if header.startswith('#'):
        header = header[1:].strip()

    columns = header.split(',')
    delta_col = None
    for idx, col in enumerate(columns):
        col_upper = col.strip().upper()
        # Match both ASCII "DELTA-TOTAL" and Unicode "\u0394 TOTAL" / "\u0394-TOTAL"
        if ('DELTA' in col_upper and 'TOTAL' in col_upper) or \
           ('\u0394' in col and 'TOTAL' in col_upper):
            delta_col = idx
            break

    if delta_col is None:
        # Try exact match
        for idx, col in enumerate(columns):
            col_s = col.strip()
            if col_s in ('DELTA TOTAL', 'DELTA-TOTAL',
                          '\u0394TOTAL', '\u0394 TOTAL', '\u0394-TOTAL',
                          'Delta Total', 'delta total'):
                delta_col = idx
                break

    if delta_col is None:
        return None

    per_frame = []
    for line in lines[1:]:
        line = line.strip()
        if not line or line.startswith('#'):
            continue
        parts = line.split(',')
        if delta_col < len(parts):
            try:
                per_frame.append(float(parts[delta_col]))
            except (ValueError, IndexError):
                pass

    if not per_frame:
        return None

    import statistics
    return {
        'per_frame_delta_total': per_frame,
        'n_frames': len(per_frame),
        'per_frame_avg': statistics.mean(per_frame),
        'per_frame_std': statistics.stdev(per_frame) if len(per_frame) > 1 else 0.0,
    }


# ── Directory scanning ──────────────────────────────────────────────────────

def scan_results(results_root: Path, n_outer_replicas: int = 3,
                 n_mmpbsa_replicas: int = 3) -> list[dict]:
    """Scan MMPBSA results directory tree and collect all entries.

    Returns list of dicts, each with:
        outer_replica, design, mmpbsa_replica, status, result_dir,
        and parsed energy data (if available).
    """
    entries = []
    for outer_rep in range(1, n_outer_replicas + 1):
        rep_dir = results_root / f"replica{outer_rep}"
        if not rep_dir.is_dir():
            continue
        for design_dir in sorted(rep_dir.iterdir()):
            if not design_dir.is_dir():
                continue
            for mmpbsa_rep in range(1, n_mmpbsa_replicas + 1):
                mmpbsa_dir = design_dir / f"mmpbsa_r{mmpbsa_rep}"
                if not mmpbsa_dir.is_dir():
                    continue

                entry = {
                    'outer_replica': outer_rep,
                    'design': design_dir.name,
                    'mmpbsa_replica': mmpbsa_rep,
                    'result_dir': str(mmpbsa_dir),
                }

                # Check completion status
                done_file = mmpbsa_dir / "DONE"
                final_file = mmpbsa_dir / "FINAL_RESULTS_MMPBSA.dat"

                if done_file.exists() and final_file.exists():
                    parsed = parse_final_results(final_file)
                    if parsed:
                        entry['status'] = 'done'
                        entry.update(parsed)
                    else:
                        entry['status'] = 'parse_error'
                elif final_file.exists():
                    parsed = parse_final_results(final_file)
                    if parsed:
                        entry['status'] = 'done_no_sentinel'
                        entry.update(parsed)
                    else:
                        entry['status'] = 'failed'
                else:
                    # Check if SLURM job is running
                    slurm_files = list(mmpbsa_dir.glob("slurm_mmpbsa_*.out"))
                    slurm_err = list(mmpbsa_dir.glob("slurm_mmpbsa_*.err"))
                    has_logs = len(slurm_files) > 0 or len(slurm_err) > 0
                    if has_logs:
                        # Check for failure keywords
                        failed = False
                        fail_keywords = [
                            "Traceback (most recent call last)",
                            "CANCELLED", "DUE TO TIME LIMIT",
                            "Segmentation fault", "Bus error",
                            "exit code 1", "ERROR:", "FATAL ERROR",
                        ]
                        for log_file in slurm_files + slurm_err:
                            try:
                                log_text = log_file.read_text(errors='replace')
                                if any(kw in log_text for kw in fail_keywords):
                                    failed = True
                                    break
                            except OSError:
                                pass
                        entry['status'] = 'failed' if failed else 'running'
                    else:
                        entry['status'] = 'pending'

                entries.append(entry)

    return entries


# ── Output ──────────────────────────────────────────────────────────────────

# Columns for the detailed CSV (one row per outer_replica × design × mmpbsa_replica)
DETAIL_COLUMNS = [
    'outer_replica', 'design', 'mmpbsa_replica', 'status',
]

# Energy columns to include (DELTA components are most important)
ENERGY_COLUMNS = [
    ('DELTA_TOTAL', 'DELTA TOTAL'),
    ('DELTA_VDWAALS', 'DELTA VDWAALS'),
    ('DELTA_EEL', 'DELTA EEL'),
    ('DELTA_PB', 'DELTA PB'),
    ('DELTA_POLAR', 'DELTA POLAR'),
    ('DELTA_NONPOLAR', 'DELTA NONPOLAR'),
    ('DELTA_GAS', 'DELTA GAS'),
    ('DELTA_SOLV', 'DELTA SOLV'),
    ('COMPLEX_TOTAL', 'COMPLEX TOTAL'),
    ('RECEPTOR_TOTAL', 'RECEPTOR TOTAL'),
    ('LIGAND_TOTAL', 'LIGAND TOTAL'),
]

PER_FRAME_EXTRA_COLUMNS = [
    'n_frames', 'per_frame_avg', 'per_frame_std',
]


def build_column_headers(include_per_frame: bool = False) -> list[str]:
    """Build ordered list of CSV column headers."""
    headers = list(DETAIL_COLUMNS)
    for key, label in ENERGY_COLUMNS:
        headers.append(f'{key}_avg')
        headers.append(f'{key}_std')
        headers.append(f'{key}_stderr')
    if include_per_frame:
        headers.extend(PER_FRAME_EXTRA_COLUMNS)
    return headers


def write_csv(entries: list[dict], outpath: Path, include_per_frame: bool = False,
              verbose: bool = True):
    """Write entries to CSV file."""
    headers = build_column_headers(include_per_frame)
    outpath.parent.mkdir(parents=True, exist_ok=True)

    with open(outpath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction='ignore',
                                restval='')
        writer.writeheader()
        for entry in entries:
            row = {}
            for h in headers:
                row[h] = entry.get(h, '')
            writer.writerow(row)

    if verbose:
        print(f"  CSV  -> {outpath}")


def write_aggregate_csv(entries: list[dict], outpath: Path,
                         verbose: bool = True):
    """Write per-design aggregated CSV (average across MMPBSA replicas).

    For each (outer_replica, design), compute mean ± SEM of DELTA_TOTAL_avg
    across all completed MMPBSA replicas.
    """
    # Group by (outer_replica, design)
    groups = defaultdict(list)
    for entry in entries:
        key = (entry.get('outer_replica', ''), entry.get('design', ''))
        if entry.get('status') in ('done', 'done_no_sentinel'):
            groups[key].append(entry)

    import statistics

    headers = [
        'outer_replica', 'design',
        'n_mmpbsa_replicas_done',
        'DELTA_TOTAL_mean', 'DELTA_TOTAL_sem',
        'DELTA_TOTAL_min', 'DELTA_TOTAL_max',
        'DELTA_VDWAALS_mean', 'DELTA_EEL_mean',
        'DELTA_PB_mean', 'DELTA_POLAR_mean',
        'DELTA_NONPOLAR_mean',
        'DELTA_GAS_mean', 'DELTA_SOLV_mean',
        'COMPLEX_TOTAL_mean', 'RECEPTOR_TOTAL_mean', 'LIGAND_TOTAL_mean',
    ]

    # Also add per-outer-replica-averaged components
    agg_rows = []
    for (outer_rep, design), group_entries in sorted(groups.items()):
        n = len(group_entries)

        # DELTA TOTAL values
        dt_avgs = [e.get('DELTA_TOTAL_avg') for e in group_entries
                    if e.get('DELTA_TOTAL_avg') is not None]
        dt_stds = [e.get('DELTA_TOTAL_std') for e in group_entries
                    if e.get('DELTA_TOTAL_std') is not None]

        if not dt_avgs:
            continue

        mean_dt = statistics.mean(dt_avgs)
        sem_dt = statistics.stdev(dt_avgs) / (n ** 0.5) if n > 1 else 0.0

        row = {
            'outer_replica': outer_rep,
            'design': design,
            'n_mmpbsa_replicas_done': n,
            'DELTA_TOTAL_mean': round(mean_dt, 4),
            'DELTA_TOTAL_sem': round(sem_dt, 4),
            'DELTA_TOTAL_min': round(min(dt_avgs), 4),
            'DELTA_TOTAL_max': round(max(dt_avgs), 4),
        }

        # Average other DELTA components
        for comp in ['VDWAALS', 'EEL', 'PB', 'POLAR', 'NONPOLAR', 'GAS', 'SOLV']:
            key = f'DELTA_{comp}_avg'
            vals = [e.get(key) for e in group_entries
                    if e.get(key) is not None]
            row[f'DELTA_{comp}_mean'] = round(statistics.mean(vals), 4) if vals else ''

        # Average COMPLEX/RECEPTOR/LIGAND TOTAL
        for prefix in ['COMPLEX', 'RECEPTOR', 'LIGAND']:
            key = f'{prefix}_TOTAL_avg'
            vals = [e.get(key) for e in group_entries
                    if e.get(key) is not None]
            row[f'{prefix}_TOTAL_mean'] = round(statistics.mean(vals), 4) if vals else ''

        agg_rows.append(row)

    # Also include entries with no completed MMPBSA replicas
    all_keys = set()
    for entry in entries:
        all_keys.add((entry.get('outer_replica', ''), entry.get('design', '')))
    for key in sorted(all_keys):
        if key not in groups:
            agg_rows.append({
                'outer_replica': key[0],
                'design': key[1],
                'n_mmpbsa_replicas_done': 0,
            })

    outpath.parent.mkdir(parents=True, exist_ok=True)
    with open(outpath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction='ignore',
                                restval='')
        writer.writeheader()
        writer.writerows(agg_rows)

    if verbose:
        print(f"  Agg  CSV -> {outpath}")


def write_excel(entries: list[dict], outpath: Path, include_per_frame: bool = False,
                aggregate_entries: list[dict] | None = None,
                agg_path: Path | None = None,
                verbose: bool = True):
    """Write Excel file with two sheets: detail and aggregate."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("WARNING: openpyxl not installed; skipping Excel output. "
              "Install with:  pip install openpyxl")
        return

    headers = build_column_headers(include_per_frame)
    outpath.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MMPBSA Detail"

    # Styles
    header_fill = PatternFill("solid", fgColor="2E86AB")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    done_fill = PatternFill("solid", fgColor="FFFFFF")
    pending_fill = PatternFill("solid", fgColor="FFF3CD")
    failed_fill = PatternFill("solid", fgColor="F8D7DA")
    running_fill = PatternFill("solid", fgColor="D4EDDA")
    thin_side = Side(style='thin', color='CCCCCC')
    thin_border = Border(left=thin_side, right=thin_side,
                         top=thin_side, bottom=thin_side)

    # Header row
    for col_i, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_i, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Data rows
    for row_i, entry in enumerate(entries, start=2):
        status = entry.get('status', '')
        if status in ('done', 'done_no_sentinel'):
            fill = done_fill
        elif status == 'running':
            fill = running_fill
        elif status == 'failed' or status == 'parse_error':
            fill = failed_fill
        else:
            fill = pending_fill

        for col_i, col_name in enumerate(headers, start=1):
            val = entry.get(col_name, '')
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Color entire row based on status
        # (Already done via fill per cell)

    # Auto column width
    for col_i, col_name in enumerate(headers, start=1):
        max_len = max(
            len(str(col_name)),
            *[len(str(e.get(col_name, ''))) for e in entries[:100]],
        )
        ws.column_dimensions[get_column_letter(col_i)].width = min(max_len + 3, 25)

    ws.freeze_panes = 'A2'
    wb.save(outpath)
    if verbose:
        print(f"  XLSX -> {outpath}")


def print_table(entries: list[dict]):
    """Print a human-readable summary table to stdout."""
    n_done = n_pending = n_failed = n_running = 0
    dt_values = []

    # Column widths
    w_rep = 8
    w_des = 40
    w_mrep = 8
    w_status = 10
    w_dt = 14
    w_dt_std = 12

    header = (
        f"{'out_rep':>{w_rep}}  "
        f"{'design':<{w_des}}  "
        f"{'mmp_rep':>{w_mrep}}  "
        f"{'status':<{w_status}}  "
        f"{'DTOTAL_avg':>{w_dt}}  "
        f"{'DT_std':>{w_dt_std}}"
    )
    print("\n" + "=" * len(header))
    print(header)
    print("=" * len(header))

    for e in entries:
        status = e.get('status', '')
        dt_avg = e.get('DELTA_TOTAL_avg', '')
        dt_std = e.get('DELTA_TOTAL_std', '')

        if status in ('done', 'done_no_sentinel'):
            n_done += 1
            dt_s = f"{dt_avg:+.3f}" if isinstance(dt_avg, (int, float)) else str(dt_avg)
            dt_std_s = f"±{dt_std:.3f}" if isinstance(dt_std, (int, float)) else str(dt_std)
            if isinstance(e.get('DELTA_TOTAL_avg'), (int, float)):
                dt_values.append(e['DELTA_TOTAL_avg'])
        elif status == 'failed' or status == 'parse_error':
            n_failed += 1
            dt_s = ''
            dt_std_s = ''
        elif status == 'running':
            n_running += 1
            dt_s = ''
            dt_std_s = ''
        else:
            n_pending += 1
            dt_s = ''
            dt_std_s = ''

        print(
            f"{e.get('outer_replica', ''):>{w_rep}}  "
            f"{e.get('design', ''):<{w_des}}  "
            f"{e.get('mmpbsa_replica', ''):>{w_mrep}}  "
            f"{status:<{w_status}}  "
            f"{dt_s:>{w_dt}}  "
            f"{dt_std_s:>{w_dt_std}}"
        )

    print("=" * len(header))
    print(f"  Total: {len(entries)}  |  done: {n_done}  |  "
          f"running: {n_running}  |  failed: {n_failed}  |  pending: {n_pending}")

    if dt_values:
        import statistics
        print(f"\n  DELTA TOTAL across {len(dt_values)} completed replicas:")
        print(f"    mean = {statistics.mean(dt_values):+.3f} kcal/mol")
        if len(dt_values) > 1:
            print(f"    std  = {statistics.stdev(dt_values):.3f} kcal/mol")
            print(f"    SEM  = {statistics.stdev(dt_values) / len(dt_values)**0.5:.3f} kcal/mol")
        print(f"    min  = {min(dt_values):+.3f}  |  max = {max(dt_values):+.3f}")
    print()


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Collect MMPBSA results into a summary CSV table.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        '--results', default=None,
        help="Path to MMPBSA results root directory "
             "(default: auto-detect as ~/MMPBSA/results)",
    )
    parser.add_argument(
        '--out', default=None,
        help="Output CSV file path (default: <results_root>/../mmpbsa_summary.csv)",
    )
    parser.add_argument(
        '--excel', action='store_true',
        help="Also write an Excel (.xlsx) file alongside the CSV",
    )
    parser.add_argument(
        '--per-frame', action='store_true',
        help="Include per-frame DELTA TOTAL data from gmx_MMPBSA_plot.csv",
    )
    parser.add_argument(
        '--aggregate', action='store_true',
        help="Write an additional aggregated CSV (mean across MMPBSA replicas per design)",
    )
    parser.add_argument(
        '--designs', nargs='+', metavar='DESIGN',
        help="Only include these design names (space-separated)",
    )
    parser.add_argument(
        '--designs-file', metavar='FILE',
        help="Read design names to include from a text file (one per line)",
    )
    parser.add_argument(
        '--outer-replicas', type=int, default=3,
        help="Number of outer replica directories to scan (default: 3)",
    )
    parser.add_argument(
        '--mmpbsa-replicas', type=int, default=3,
        help="Number of MMPBSA replicas per design (default: 3)",
    )
    parser.add_argument(
        '--no-print', action='store_true',
        help="Suppress the console table printout",
    )
    parser.add_argument(
        '--debug', action='store_true',
        help="Print debugging info (e.g., first parsed file details)",
    )
    args = parser.parse_args()

    # ── Resolve results root ──────────────────────────────────────────────
    if args.results:
        results_root = Path(args.results).expanduser().resolve()
    else:
        hpc_path = Path.home() / "MMPBSA" / "results"
        script_rel = Path(__file__).parent.parent / "results"
        if hpc_path.is_dir():
            results_root = hpc_path
        elif script_rel.is_dir():
            results_root = script_rel.resolve()
        else:
            print(
                "ERROR: Could not auto-detect results directory.\n"
                "  Pass --results /path/to/MMPBSA/results  explicitly.",
                file=sys.stderr,
            )
            sys.exit(1)

    print(f"Scanning: {results_root}")
    if not results_root.is_dir():
        print(f"ERROR: results directory does not exist: {results_root}",
              file=sys.stderr)
        sys.exit(1)

    # ── Design filter ─────────────────────────────────────────────────────
    design_filter: set | None = None
    names: list[str] = []
    if args.designs:
        names.extend(args.designs)
    if args.designs_file:
        fpath = Path(args.designs_file)
        if not fpath.is_file():
            print(f"ERROR: designs file not found: {fpath}", file=sys.stderr)
            sys.exit(1)
        names.extend(
            line.strip() for line in fpath.read_text().splitlines()
            if line.strip() and not line.startswith('#')
        )
    if names:
        design_filter = set(names)
        print(f"Filter: {len(design_filter)} design(s) specified")

    # ── Scan and collect ──────────────────────────────────────────────────
    entries = scan_results(results_root, n_outer_replicas=args.outer_replicas,
                           n_mmpbsa_replicas=args.mmpbsa_replicas)

    # Apply design filter
    if design_filter:
        entries = [e for e in entries if e.get('design', '') in design_filter]

    if not entries:
        print("No MMPBSA result directories found. Check --results path or --designs filter.")
        sys.exit(0)

    # ── Debug: show first parsed result ────────────────────────────────────
    if args.debug:
        for entry in entries:
            if entry.get('status') in ('done', 'done_no_sentinel', 'parse_error'):
                dbg_file = Path(entry['result_dir']) / "FINAL_RESULTS_MMPBSA.dat"
                if dbg_file.is_file():
                    print(f"\n{'='*70}")
                    print(f"DEBUG: Parsing {dbg_file}")
                    parsed = parse_final_results(dbg_file)
                    if parsed:
                        print(f"  Parsed keys ({len(parsed)}):")
                        for k in sorted(parsed.keys()):
                            print(f"    {k} = {parsed[k]}")
                    else:
                        print("  FAILED to parse. First 30 lines:")
                        try:
                            lines = dbg_file.read_text(encoding='utf-8', errors='replace').splitlines()
                            for i, line in enumerate(lines[:30]):
                                print(f"    {i+1}: {repr(line)}")
                        except OSError:
                            print("  (cannot read file)")
                    print(f"{'='*70}\n")
                    break
        else:
            print("DEBUG: No parseable/failed results found.")

    # ── Per-frame data (optional) ─────────────────────────────────────────
    if args.per_frame:
        print("Parsing per-frame CSV data...")
        for entry in entries:
            if entry.get('status') in ('done', 'done_no_sentinel'):
                csv_path = Path(entry['result_dir']) / "gmx_MMPBSA_plot.csv"
                pf_data = parse_per_frame_csv(csv_path)
                if pf_data:
                    entry['n_frames'] = pf_data['n_frames']
                    entry['per_frame_avg'] = round(pf_data['per_frame_avg'], 4)
                    entry['per_frame_std'] = round(pf_data['per_frame_std'], 4)

    # ── Print summary table ──────────────────────────────────────────────
    if not args.no_print:
        print_table(entries)

    # ── Write outputs ────────────────────────────────────────────────────
    if args.out:
        csv_path = Path(args.out).expanduser().resolve()
    else:
        csv_path = results_root.parent / "mmpbsa_summary.csv"

    write_csv(entries, csv_path, include_per_frame=args.per_frame)

    if args.aggregate:
        agg_path = csv_path.with_name("mmpbsa_aggregate.csv")
        write_aggregate_csv(entries, agg_path)

    if args.excel:
        xlsx_path = csv_path.with_suffix('.xlsx')
        write_excel(entries, xlsx_path, include_per_frame=args.per_frame)

    print("Done.")


if __name__ == '__main__':
    main()